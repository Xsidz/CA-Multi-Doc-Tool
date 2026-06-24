"""Excel export service — builds a styled .xlsx workbook from parsed rows."""
from __future__ import annotations

import io
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Styling constants ─────────────────────────────────────────────────────────
# Header: dark navy fill (#1E3A5F), white bold text
HEADER_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Alternating row fills: even rows white, odd rows light blue-grey (#F8FAFC)
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
ALT_FILL   = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
DATA_FONT  = Font(name="Calibri", size=10)
DATA_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=False)

# Column width caps
_MIN_COL_WIDTH = 8
_MAX_COL_WIDTH = 50


def build_excel(doc_type: str, rows: list[dict[str, Any]]) -> bytes:
    """Build a styled Excel workbook from *rows* and return raw bytes.

    Layout:
    - Sheet tab named after *doc_type* (uppercase, max 31 chars per Excel spec)
    - Row 1: header row with navy background, white bold font, center-aligned
    - Rows 2+: data rows with alternating white / #F8FAFC fills
    - Freeze pane set at A2 so the header stays visible while scrolling
    - Column widths auto-fitted based on the longest cell value (max 50 chars)

    Returns an empty (but valid) workbook if *rows* is empty.
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    # Excel sheet names are limited to 31 characters
    sheet_name = doc_type.upper()[:31]
    ws.title = sheet_name

    if not rows:
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    # Derive headers from the first row; preserve insertion order
    headers: list[str] = list(rows[0].keys())
    num_cols = len(headers)

    # ── Header row (row 1) ───────────────────────────────────────────────────
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN

    ws.row_dimensions[1].height = 30

    # ── Data rows (rows 2 … N+1) ─────────────────────────────────────────────
    for row_idx, row_data in enumerate(rows, start=2):
        # Alternate: even data rows (row_idx 2, 4, …) white; odd rows #F8FAFC
        fill = WHITE_FILL if row_idx % 2 == 0 else ALT_FILL
        for col_idx, header in enumerate(headers, start=1):
            raw_value = row_data.get(header, "")
            # Convert None → empty string for cleaner display
            value = "" if raw_value is None else raw_value
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGN

    # ── Auto-fit column widths ────────────────────────────────────────────────
    # Estimate width as max visible character count across header + all data cells
    for col_idx in range(1, num_cols + 1):
        col_letter = get_column_letter(col_idx)

        # Start with the header length
        max_len = len(str(ws.cell(row=1, column=col_idx).value or ""))

        for row_idx in range(2, len(rows) + 2):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is not None:
                # For numbers, format with commas to match display length
                cell_len = len(str(cell_val))
                max_len = max(max_len, cell_len)

        adjusted_width = max(_MIN_COL_WIDTH, min(max_len + 4, _MAX_COL_WIDTH))
        ws.column_dimensions[col_letter].width = adjusted_width

    # ── Freeze header row ─────────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
